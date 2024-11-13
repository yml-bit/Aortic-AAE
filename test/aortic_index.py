import SimpleITK as sitk
import os
import numpy as np
from skimage import measure, morphology
from scipy.spatial.distance import cdist
import h5py
import matplotlib.pyplot as plt
from skimage.morphology import binary_dilation
import copy
from mayavi import mlab
import cv2
from skimage.morphology import binary_closing
from scipy.stats import pearsonr,binned_statistic
import pingouin as pg
from sklearn.metrics import mean_absolute_error,r2_score,mean_absolute_percentage_error
from natsort import natsorted
import openpyxl
from sklearn.metrics import confusion_matrix,roc_curve, auc,cohen_kappa_score, matthews_corrcoef
import matplotlib.pyplot as plt
from sklearn.preprocessing import label_binarize
import pandas as pd
from scipy.ndimage import label, generate_binary_structure
# 设置打印选项,使得所有数组都以小数形式输出,且设置小数点后保留的位数
np.set_printoptions(suppress=True, precision=8)  # suppress=True 禁用科学记数法,precision设置小数点后的位数
from skimage.measure import regionprops
from collections import Counter
import heapq
import statsmodels.api as sm

#remain the bigest Connected region
def remove_regions(mask):
    mask1=copy.deepcopy(mask)
    mask1=np.where(mask1>0,1,0)
    # mask2 = mask
    # mask2=np.where(mask2>1,1,0)
    segmentation_sitk = sitk.GetImageFromArray(mask1)
    # 计算标签图像的连通分量
    connected_components_filter = sitk.ConnectedComponentImageFilter()
    labeled_image = connected_components_filter.Execute(segmentation_sitk)

    # 获取每个连通分量的大小
    label_shape_filter = sitk.LabelShapeStatisticsImageFilter()
    label_shape_filter.Execute(labeled_image)

    # 找到最大的连通分量ID
    max_size = 0
    largest_label = 0
    for i in range(1, label_shape_filter.GetNumberOfLabels() + 1):  # Label index starts from 1
        if label_shape_filter.GetNumberOfPixels(i) > max_size:
            max_size = label_shape_filter.GetNumberOfPixels(i)
            largest_label = i

    # 仅保留最大连通分量
    binary_mask = sitk.Equal(labeled_image, largest_label)
    cleaned_segmentation = sitk.Cast(binary_mask, segmentation_sitk.GetPixelID())
    cleaned_segmentation = sitk.GetArrayFromImage(cleaned_segmentation)
    cleaned_segmentation=cleaned_segmentation*mask
    # print(cleaned_segmentation.max())
    return cleaned_segmentation.astype(np.int16)

##remain the  Connected region whichs more than 1000 voxel
def remove_small_volums(mask):
    mask1 = np.where(mask > 0, 1, 0)
    segmentation_sitk = sitk.GetImageFromArray(mask1)
    # 计算标签图像的连通分量
    connected_components_filter = sitk.ConnectedComponentImageFilter()
    labeled_image = connected_components_filter.Execute(segmentation_sitk)
    # 获取每个连通分量的大小
    label_shape_filter = sitk.LabelShapeStatisticsImageFilter()
    label_shape_filter.Execute(labeled_image)
    # 初始化一个空的数组来存储处理后的mask
    cleaned_segmentation = np.zeros_like(mask1)
    # 遍历每个连通分量,保留体积大于min_volume的连通区域
    for i in range(1, label_shape_filter.GetNumberOfLabels() + 1):
        if label_shape_filter.GetNumberOfPixels(i) >= 14000:
            binary_mask = sitk.Equal(labeled_image, i)
            binary_mask_array = sitk.GetArrayFromImage(binary_mask)
            cleaned_segmentation[binary_mask_array == 1] = 1
    # 返回处理后的mask
    cleaned_segmentation = cleaned_segmentation * mask
    return cleaned_segmentation.astype(np.int16)

def get_longest_3d_mask(mask):
    mask1=copy.deepcopy(mask)
    mask=np.where(mask>0,1,0)
    # Step 1: Extract connected components
    labeled_mask, num_features = label(mask)

    if num_features == 0:
        return None  # No connected components found

    # Step 2: Compute geometric features and select the longest component
    max_length = 0
    longest_component = None
    for i in range(1, num_features + 1):
        component_mask = (labeled_mask == i).astype(np.uint8)

        # Calculate the length using the main axis of the bounding box
        props = regionprops(component_mask)
        if len(props) > 0:
            bbox = props[0].bbox
            # Assuming the bounding box is sorted in ascending order
            length = max(bbox[3] - bbox[0], bbox[4] - bbox[1], bbox[5] - bbox[2])

            # Alternatively, you can calculate the distance between the centers of mass
            # to estimate the length as the longest diagonal
            # center_of_masses = [center_of_mass(component_mask)]
            # length = max([euclidean(cm1, cm2) for cm1, cm2 in zip(center_of_masses[:-1], center_of_masses[1:])])

            if length > max_length:
                max_length = length
                longest_component = component_mask

    return mask1*longest_component

def my_mape(y_true, y_pred):
    if y_true==0 and y_pred!=0:
        error=100
    elif y_pred==0 and y_true!=0:
        error=0
    elif y_pred==0 and y_true==0:
        error=0
    else:
        error = np.abs((y_true - y_pred) / y_true) * 100.0
        error = np.clip(error, 0, 100)
    return error

# def remove_ascend(image_stack):
#     # 定义结构元素用于确定连通性（这里假定是8邻域）
#     struct = generate_binary_structure(2, 2)
#
#     # 对每个切片进行连通组件分析
#     ascend_index=[]
#     i=0
#     labeled_slices = []
#     for slice_lume in image_stack:
#         img_label, num = measure.label(slice_lume, connectivity=2, return_num=True)
#         props = measure.regionprops(img_label)
#         props_sorted = sorted(props, key=lambda x: x.area, reverse=True)
#         if len(props_sorted) > 1:
#             if props_sorted[1].area>600:
#                 a=1
#
#         if len(props_sorted) > 1 and props_sorted[0].area > 600:
#             # max_label = props_sorted[1].label
#             if props_sorted[0].area < 3 * (props_sorted[1].area):
#                 max_label = props_sorted[1].label
#             else:
#                 max_label = props_sorted[0].label
#         else:
#             max_label = props_sorted[0].label
#         i=i+1
#     # 返回过滤后的3D mask
#     return np.stack(labeled_slices)

def caw(mask):
    """
    使用矢量化操作重新赋值3D掩模中的像素值。

    :param mask: 一个三维numpy数组,代表3D掩模。
    :return: 重新赋值后的3D掩模。
    """
    reassigned_mask = np.zeros_like(mask)

    # 数值区间及其对应的赋值
    reassigned_mask[(mask >= 130) & (mask <= 199)] = 1
    reassigned_mask[(mask >= 200) & (mask <= 299)] = 2
    reassigned_mask[(mask >= 300) & (mask <= 399)] = 3
    reassigned_mask[mask >= 400] = 4

    return reassigned_mask

def diameter_area(mask,ncct_arrayy):
    # lumen_mask = np.where(mask == 1, 1, 0)  # 管腔（正常血液流动区域）
    v1=copy.deepcopy(mask)
    z_dims = mask.shape[0]
    area_per_lumen = []
    diameter_per_lumen = []

    for slice_idx in range(z_dims):  # 如果存在升主动脉,去掉升主动脉
        slice_idx=z_dims-slice_idx-1
        slice_lume = mask[slice_idx, :, :]
        img_label, num = measure.label(slice_lume, connectivity=2, return_num=True)#填充空洞
        props = measure.regionprops(img_label)
        props_sorted = sorted(props, key=lambda x: x.area, reverse=True)
        if len(props_sorted) > 1 and props_sorted[0].area > 300 and slice_idx > 0:
            if props_sorted[1].area>300:
                v1[slice_idx+1:, :, :]=0
                v1=remove_regions(v1)
                v1[slice_idx+1:, :, :]=mask[slice_idx+1:, :, :]
                mask=copy.deepcopy(v1)
                continue#处理完就结束,节约时间
    ncct_arrayyy=ncct_arrayy*mask
    sum_z_axis = np.sum(ncct_arrayyy, axis=(1, 2))

    flag_area=np.ones([512,512])
    for slice_idx in range(z_dims):  # 获取当前切片
        slice_lume = mask[slice_idx, :, :]
        if slice_lume.sum()==0:
            diameter_per_lumen.append(0)
            area_per_lumen.append(0)  # 将面积添加到列表中
        else:
            img_label, num = measure.label(slice_lume, connectivity=2, return_num=True)
            props = measure.regionprops(img_label)
            props_sorted = sorted(props, key=lambda x: x.area, reverse=True)
            max_label = props_sorted[0].label

            if len(props_sorted) > 1 and props_sorted[1].area>300:
                if props_sorted[0].area<2*props_sorted[1].area:
                    max_label0 = props_sorted[0].label
                    slice_lume0 = (img_label == max_label0).astype(int)
                    overlap0 = np.logical_and(slice_lume0, flag_area)
                    overlap_area0 = np.count_nonzero(overlap0)

                    max_label = props_sorted[1].label
                    slice_lume = (img_label == max_label).astype(int)
                    overlap = np.logical_and(slice_lume, flag_area)
                    overlap_area = np.count_nonzero(overlap)
                    if overlap_area>overlap_area0:
                        max_label = props_sorted[1].label
                    else:
                        max_label = props_sorted[0].label

            slice_lume = (img_label == max_label).astype(int)
            flag_area=slice_lume
            filled_slice_lume = binary_closing(slice_lume)  #空洞填充,以便于计算直径
            gray_img = np.uint8(filled_slice_lume * 255)
            contours, _ = cv2.findContours(gray_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            dist_map = cv2.distanceTransform(gray_img, cv2.DIST_L2, cv2.DIST_MASK_PRECISE)
            _, radius, _, center = cv2.minMaxLoc(dist_map)
            diameter = 2 * radius
            # threshold = np.percentile(distances, 2)
            area = np.sum(slice_lume > 0)  # 计算填充前的面积
            if area > np.pi * (diameter / 2) ** 2:
                area = np.pi * diameter ** 2
            diameter_per_lumen.append(diameter)
            area_per_lumen.append(area)

    return diameter_per_lumen, area_per_lumen,sum_z_axis

def caculate_index(path,ncct_path, path_save):
    # 一：读取保存有主动脉mask的NIfTI文件
    ncct = sitk.ReadImage(ncct_path, sitk.sitkInt16)
    ncct_array = sitk.GetArrayFromImage(ncct)
    read = sitk.ReadImage(path, sitk.sitkInt16)  # 使用sitk重新保存,这样占用内存小很多
    img_array = sitk.GetArrayFromImage(read)
    img_array[img_array>2]=3
    # img_array = remove_small_volums(img_array)#去掉冗余  remove_small_volums
    # img_array = remove_regions(img_array)  # 去掉冗余  remove_small_volums
    img_array=get_longest_3d_mask(img_array)
    mask_image=copy.deepcopy(img_array)
    out = sitk.GetImageFromArray(img_array)
    sitk.WriteImage(out, path_save)#之前已经执行过了

    Ca = copy.deepcopy(mask_image)  # 钙化
    reassigned_Ca = np.where(Ca == 2, 1, 0)
    ncct_arrayy=ncct_array*reassigned_Ca
    ncct_arrayy=caw(ncct_arrayy)

    lumen_mask=copy.deepcopy(mask_image)#管腔
    lumen_mask[lumen_mask > 1] = 0
    lumen_mask = remove_small_volums(lumen_mask)
    diameter_per_lumen,area_per_lumen,_=diameter_area(lumen_mask,ncct_arrayy)

    total_mask = np.where(mask_image >0, 1, 0)#血管
    total_mask=remove_small_volums(total_mask)
    diameter_per_total,area_per_total,ncct_arrayyy=diameter_area(total_mask,ncct_arrayy)

    # per_index = diameter_per_lumen / diameter_per_total  # 钙化指数应该是钙化面积/(管腔+钙化)
    # per_index = np.ones(mask_image.shape[0]) * 0
    per_index = [diameter_per_lumen[i] / diameter_per_total[i] if diameter_per_total[i] != 0 else 0 for i in range(len(diameter_per_total))]
    per_index= np.array(per_index)
    per_index[per_index==0]=1
    per_index=1-per_index
    # per_index[per_index == 1] = 0
    per_index = np.clip(per_index, 0, 1)
    per_index[np.isnan(per_index)] = 0
    per_index[np.isinf(per_index)] = 0
    per_index = per_index.tolist()
    # per_index = [num for num in per_index if num != 0]

    lumen_mask=copy.deepcopy(mask_image)#管腔
    lumen_mask[lumen_mask > 2] = 1 #计算钙化斑块引起的狭窄
    lumen_mask[lumen_mask > 1] = 0
    #lumen_mask = remove_small_volums(lumen_mask)
    diameter_per_lumenc,area_per_lumenc,_=diameter_area(lumen_mask,ncct_arrayy)
    per_indexc = [diameter_per_lumenc[i] / diameter_per_total[i] if diameter_per_total[i] != 0 else 0 for i in range(len(diameter_per_total))]
    per_indexc= np.array(per_indexc)
    per_indexc[per_indexc==0]=1
    per_indexc=1-per_indexc
    # per_indexc[per_indexc == 1] = 0
    per_indexc = np.clip(per_indexc, 0, 1)
    per_indexc[np.isnan(per_indexc)] = 0
    per_indexc[np.isinf(per_indexc)] = 0
    per_indexc = per_indexc.tolist()

    lumen_mask=copy.deepcopy(mask_image)#管腔
    lumen_mask[lumen_mask > 2] = 0 #计算软斑块引起的狭窄
    lumen_mask[lumen_mask > 0] = 1
    lumen_mask = remove_small_volums(lumen_mask)
    diameter_per_lumens,area_per_lumens,_=diameter_area(lumen_mask,ncct_arrayy)

    # per_index = diameter_per_lumen / diameter_per_total  # 钙化指数应该是钙化面积/(管腔+钙化)
    # per_index = np.ones(mask_image.shape[0]) * 0
    per_indexs = [diameter_per_lumens[i] / diameter_per_total[i] if diameter_per_total[i] != 0 else 0 for i in range(len(diameter_per_total))]
    per_indexs= np.array(per_indexs)
    per_indexs[per_indexs==0]=1
    per_indexs=1-per_indexs
    # per_indexs[per_indexs == 1] = 0
    per_indexs = np.clip(per_indexs, 0, 1)
    per_indexs[np.isnan(per_indexs)] = 0
    per_indexs[np.isinf(per_indexs)] = 0
    per_indexs = per_indexs.tolist()

    # file=path.replace(".nii.gz",'.h5')
    h5_path_save = path_save.replace(".nii.gz", ".h5")
    if os.path.exists(h5_path_save):
        os.remove(h5_path_save)

    with h5py.File(h5_path_save, 'w') as f:# 创建一个dataset
        f.create_dataset('diameter_per_lumen', data=diameter_per_lumen)#管腔
        f.create_dataset('diameter_per_lumenc', data=diameter_per_lumenc)#管腔
        f.create_dataset('diameter_per_lumens', data=diameter_per_lumens)#管腔
        f.create_dataset('area_per_lumen', data=area_per_lumen)
        f.create_dataset('area_per_lumecn', data=area_per_lumenc)
        f.create_dataset('area_per_lumens', data=area_per_lumens)
        f.create_dataset('diameter_per_total', data=diameter_per_total)#整体
        f.create_dataset('area_per_total', data=area_per_total)
        f.create_dataset('per_index', data=per_index)#钙化 ncct_arrayyy
        f.create_dataset('per_indexc', data=per_indexc)#钙化 ncct_arrayyy
        f.create_dataset('per_indexs', data=per_indexs)  # 软斑块引起的狭窄
        f.create_dataset('ncct_arrayyy', data=ncct_arrayyy)

#已将这部分整合到计算模块代码中
def Aortic_index_internal():
    path = "/media/bit301/data/yml/data/p2_nii/internal/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "2.nii.gz" in path:
                path_list.append(path)
    i = 0

    for path in path_list:
        # if i<190:
        #     i=i+1
        #     continue
        # path = "/media/bit301/data/yml/data/p2_nii/external/cq/dis/xz/PA247/2.nii.gz"
        path_save = path.replace("internal", "statisticians/internal")  # lo:97  x1:54  xlo2:10
        out_put = os.path.dirname(path_save)
        if not os.path.isdir(out_put):
            os.makedirs(out_put)
        ncct_path=path.replace("2.nii.gz","0.nii.gz" )
        caculate_index(path, ncct_path, path_save)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)

def Aortic_index_external():
    # path = "/media/bit301/data/yml/data/p2_nii/internal/"
    path = "/media/bit301/data/yml/data/p2_nii/external/cq/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "2.nii.gz" in path:
                path_list.append(path)
    i = 0

    # path_list = []
    # path_list=["/media/bit301/data/yml/data/p2_nii/external/lz/dis/xz/PA171/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/lz/dis/dmzyyh/PA36/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/lz/dis/dmzyyh/PA53/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/lz/dis/dmzyyh/PA95/2.nii.gz"]

    # path_list=["/media/bit301/data/yml/data/p2_nii/external/cq/dis/dmzyyh/PA28/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/cq/dis/xz/PA247/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/cq/dis/dmzyyh/PA171/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/cq/dis/dmzyyh/PA159/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/cq/dis/dmzyyh/PA91/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/cq/dis/dmzyyh/PA70/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/cq/dis/dmzyyh/PA63/2.nii.gz",
    #            "/media/bit301/data/yml/data/p2_nii/external/cq/nor/PA280/2.nii.gz"]

    for path in path_list:
        # if i<190:
        #     i=i+1
        #     continue
        # path = "/media/bit301/data/yml/data/p2_nii/external/cq/dis/xz/PA247/2.nii.gz"
        path_save = path.replace("external", "statisticians/Aortic_index")  # lo:97  x1:54  xlo2:10
        # path_save = path.replace("internal", "Aortic_indexx")  # lo:97  x1:54  xlo2:10
        out_put = path_save.split("2.nii.gz")[0]
        if not os.path.isdir(out_put):
            os.makedirs(out_put)
        path_ncct=path.replace("2.nii.gz","0.nii.gz" )
        caculate_index(path, path_ncct, path_save)
        pathh=path.replace("external", "test/MedNeXtx22")
        path_savee = path_save.replace("2.nii.gz", "22.nii.gz")
        caculate_index(pathh,path_ncct, path_savee)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)

#基于之前的钙化做进一步统计。已将这部分整合到计算模块代码中
def post_process():
    path = "/media/bit301/data/yml/data/p2_nii/Aortic_index/" #cq
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "/2.h5" in path:
                path_list.append(path)
    path_list = natsorted(path_list)
    ii = 1
    for path in path_list:
        with h5py.File(path, 'r') as f_gd:#评估四分位数据
            diameter_per_lumen = f_gd['diameter_per_lumen'][:]
            area_per_lumen = f_gd['area_per_lumen'][:]
            diameter_per_total = f_gd['diameter_per_total'][:]
            area_per_total = f_gd['area_per_total'][:]
            per_index = f_gd['per_index'][:]
            ncct_arrayyy = f_gd['ncct_arrayyy'][:]
        ca_gd = np.sum(ncct_arrayyy, axis=(1, 2))
        with h5py.File(path, 'w') as f:  # 创建一个dataset
            f.create_dataset('diameter_per_lumen', data=diameter_per_lumen)  # 管腔
            f.create_dataset('area_per_lumen', data=area_per_lumen)
            f.create_dataset('diameter_per_total', data=diameter_per_total)  # 整体
            f.create_dataset('area_per_total', data=area_per_total)
            f.create_dataset('per_index', data=per_index)  # 钙化 ncct_arrayyy
            f.create_dataset('ncct_arrayyy', data=ca_gd)

        path_mea=path.replace("2.h5","22.h5")
        with h5py.File(path_mea, 'r') as f_mea:#评估四分位数据
            diameter_per_lumen = f_mea['diameter_per_lumen'][:]
            area_per_lumen = f_mea['area_per_lumen'][:]
            diameter_per_total = f_mea['diameter_per_total'][:]
            area_per_total = f_mea['area_per_total'][:]
            per_index = f_mea['per_index'][:]
            ncct_arrayyy = f_mea['ncct_arrayyy'][:]
        ca_mea = np.sum(ncct_arrayyy, axis=(1, 2))
        with h5py.File(path_mea, 'w') as f:  # 创建一个dataset
            f.create_dataset('diameter_per_lumen', data=diameter_per_lumen)  # 管腔
            f.create_dataset('area_per_lumen', data=area_per_lumen)
            f.create_dataset('diameter_per_total', data=diameter_per_total)  # 整体
            f.create_dataset('area_per_total', data=area_per_total)
            f.create_dataset('per_index', data=per_index)  # 钙化 ncct_arrayyy
            f.create_dataset('ncct_arrayyy', data=ca_mea)
        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)
    print("finished!")

def aortic_index_xs(path,path_h5):
    # 一：读取保存有主动脉mask的NIfTI文件
    # ncct = sitk.ReadImage(pp, sitk.sitkInt16)
    # ncct_array = sitk.GetArrayFromImage(ncct)
    if not os.path.exists(path_h5):
        print(f"File not found: {path_h5}")
        return
    with h5py.File(path_h5, 'r') as f_gd:#评估四分位数据
        diameter_per_lumen = f_gd['diameter_per_lumen'][:]
        area_per_lumen = f_gd['area_per_lumen'][:]

        diameter_per_lumens = f_gd['diameter_per_lumens'][:]
        area_per_lumens = f_gd['area_per_lumens'][:]

        diameter_per_total = f_gd['diameter_per_total'][:]
        area_per_total = f_gd['area_per_total'][:]
        per_index = f_gd['per_index'][:]

        per_indexs = f_gd['per_indexs'][:]
        ncct_arrayyy = f_gd['ncct_arrayyy'][:]

    read = sitk.ReadImage(path, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
    img_array = sitk.GetArrayFromImage(read)
    img_array[img_array>2]=3
    mask_image=copy.deepcopy(img_array)

    lumen_mask=copy.deepcopy(mask_image)#管腔
    lumen_mask[lumen_mask > 2] = 1 #计算钙化斑块引起的狭窄
    lumen_mask[lumen_mask > 1] = 0
    lumen_mask = remove_small_volums(lumen_mask)
    diameter_per_lumenc,area_per_lumenc=diameter_area(lumen_mask)#
    per_indexc = [diameter_per_lumenc[i] / diameter_per_total[i] if diameter_per_total[i] != 0 else 0 for i in range(len(diameter_per_total))]
    per_indexc= np.array(per_indexc)
    per_indexc[per_indexc==0]=1
    per_indexc=1-per_indexc
    # per_indexc[per_indexc == 1] = 0
    per_indexc = np.clip(per_indexc, 0, 1)
    per_indexc[np.isnan(per_indexc)] = 0
    per_indexc[np.isinf(per_indexc)] = 0
    per_indexc = per_indexc.tolist()

    # lumen_mask=copy.deepcopy(mask_image)#管腔
    # lumen_mask[lumen_mask > 2] = 0 #计算软斑块引起的狭窄
    # lumen_mask[lumen_mask > 0] = 1
    # lumen_mask = remove_small_volums(lumen_mask)
    # diameter_per_lumens,area_per_lumens=diameter_area(lumen_mask)#
    # per_indexs = [diameter_per_lumens[i] / diameter_per_total[i] if diameter_per_total[i] != 0 else 0 for i in range(len(diameter_per_total))]
    # per_indexs= np.array(per_indexs)
    # per_indexs[per_indexs==0]=1
    # per_indexs=1-per_indexs
    # # per_indexs[per_indexs == 1] = 0
    # per_indexs = np.clip(per_indexs, 0, 1)
    # per_indexs[np.isnan(per_indexs)] = 0
    # per_indexs[np.isinf(per_indexs)] = 0
    # per_indexs = per_indexs.tolist()
    # per_index = [num for num in per_index if num != 0]

    if os.path.exists(path_h5):
        os.remove(path_h5)

    with h5py.File(path_h5, 'w') as f:# 创建一个dataset
        f.create_dataset('diameter_per_lumen', data=diameter_per_lumen)#管腔
        f.create_dataset('diameter_per_lumenc', data=diameter_per_lumenc)#管腔
        f.create_dataset('diameter_per_lumens', data=diameter_per_lumens)#管腔
        f.create_dataset('area_per_lumen', data=area_per_lumen)
        f.create_dataset('area_per_lumenc', data=area_per_lumenc)
        f.create_dataset('area_per_lumens', data=area_per_lumens)
        f.create_dataset('diameter_per_total', data=diameter_per_total)#整体
        f.create_dataset('area_per_total', data=area_per_total)
        f.create_dataset('per_index', data=per_index)#钙化 ncct_arrayyy
        f.create_dataset('per_indexc', data=per_indexc)  # 软斑块引起的狭窄
        f.create_dataset('per_indexs', data=per_indexs)  # 软斑块引起的狭窄
        f.create_dataset('ncct_arrayyy', data=ncct_arrayyy)

#计算软斑块的导致的狭窄情况，基于之前的计算结果
def Aortic_index_internal_xcs():
    path = "/media/bit301/data/yml/data/p2_nii/statisticians/internal/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "/2.nii.gz" in path:
                path_list.append(path)

    # f = open("/media/bit301/data/yml/project/python39/p2/Aorta_net/data/catchin.txt")  # hnnk test.txt
    # for line in f.readlines():  # tile_step_size=0.75较好处理官腔错位问题
    #     path=line.split('\n')[0].replace("/2.h5","/2.nii.gz")
    #     path_list.append(path)

    i = 0
    for path in path_list:
        # if i<190:
        #     i=i+1
        #     continue
        path_h5=path.replace("2.nii.gz", "2.h5")
        aortic_index_xs(path,path_h5)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)

#计算软斑块的导致的狭窄情况，基于之前的计算结果
def Aortic_index_external_xcs():
    f = open("/media/bit301/data/yml/project/python39/p2/Aorta_net/data/discard.txt")  # hnnk test.txt
    dis_list = []
    for line in f.readlines():  # tile_step_size=0.75较好处理官腔错位问题
        path=line.split('\n')[0].replace("0.nii.gz","2.nii.gz").split("external/")[1]
        dis_list.append(path)

    # path = "/media/bit301/data/yml/data/p2_nii/internal/"
    path = "/media/bit301/data/yml/data/p2_nii/statisticians/Aortic_indexcs/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "/2.nii.gz" in path:
                pp = path.split("Aortic_indexcs/")[1]
                if pp in dis_list:
                    continue
                path_list.append(path)
    i = 0
    for path in path_list:
        # if i<190:
        #     i=i+1
        #     continue
        # path = "/media/bit301/data/yml/data/p2_nii/external/cq/dis/xz/PA247/2.nii.gz"
        path_h5=path.replace("2.nii.gz", "2.h5")
        aortic_index_xs(path,path_h5)

        path= path.replace("2.nii.gz", "22.nii.gz")
        path_h5=path.replace("22.nii.gz", "22.h5")
        aortic_index_xs(path,path_h5)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)

def compute_ccc(x, y,rho):
    '''
    计算一致性相关系数（Concordance Correlation Coefficient, CCC）
    '''
    # rho, _ = pearsonr(x, y)
    mean_x = np.mean(x)
    mean_y = np.mean(y)
    var_x = np.var(x)
    var_y = np.var(y)
    # 计算CCC系数
    ccc = 2 * rho * np.sqrt(var_x) * np.sqrt(var_y) / (var_x + var_y + (mean_x - mean_y) ** 2) + (mean_x - mean_y) ** 2 / (
                var_x + var_y + (mean_x - mean_y) ** 2)
    return ccc

def metric1(mea,gd):
    measurement_values = np.nan_to_num(mea)  # 将NaN替换为0,Inf替换为最大或最小浮点数
    groundtruth_values = np.nan_to_num(gd)
    mae = mean_absolute_error(groundtruth_values, measurement_values)# 计算MAE
    mre=my_mape(np.nanmean(groundtruth_values), np.nanmean(measurement_values))
    if mae==0:
        pearson_corr=1
        ccc=1
    else:
        # variance = np.var(measurement_values - groundtruth_values)# 计算方差
        # mape=mean_absolute_percentage_error(groundtruth_values, measurement_values)
        # r2=r2_score(measurement_values, groundtruth_values)
        if np.all(measurement_values[0] == measurement_values) and np.all(groundtruth_values[0] == groundtruth_values):
            pearson_corr=1
        else:
            if np.all(measurement_values[0] == measurement_values):
                measurement_values=np.where(groundtruth_values>0,1,0)*1e-8
            elif np.all(groundtruth_values[0] == groundtruth_values):
                groundtruth_values=np.where(measurement_values>0,1,0)*1e-8
            pearson_corr, _ = pearsonr(measurement_values, groundtruth_values)# 计算Pearson相关系数
        ccc = compute_ccc(groundtruth_values, measurement_values,pearson_corr)#不达预期

    # print("MAE: ", mae)
    # print("Pearson Correlation Coefficient: ", pearson_corr)
    # print("CCC: ", ccc)
    # print("\b")
    return mae,mre,pearson_corr,ccc

def metric(mea,gd):
    #The distal abdominal aorta (i.e., the portion near the bifurcation that divides the right and left iliac arteries)
    # has a diameter of approximately 1.5 cm to 2.0 cm (or 15 mm to 20 mm)
    if mea.size == 0 or gd.size == 0:
        raise ValueError("Found array with 0 sample(s); at least one sample is required.")

    measurement_values = np.nan_to_num(mea)  # 将NaN替换为0,Inf替换为最大或最小浮点数
    groundtruth_values = np.nan_to_num(gd)
    mae = mean_absolute_error(groundtruth_values, measurement_values)# 计算MAE
    # mapes,_=my_mape1(groundtruth_values, measurement_values)
    # mape=np.mean(mapes[start_index:end_index])
    mape=my_mape(np.nanmean(gd),np.nanmean(mea))
    if mae==0:
        pearson_corr=1
        ccc=1
    else:
        # variance = np.var(measurement_values - groundtruth_values)# 计算方差
        # mape=mean_absolute_percentage_error(groundtruth_values, measurement_values)
        # r2=r2_score(measurement_values, groundtruth_values)
        if np.all(measurement_values[0] == measurement_values) or np.all(groundtruth_values[0] == groundtruth_values):
            pearson_corr=1
        else:
            if np.all(measurement_values[0] == measurement_values):
                measurement_values=np.where(groundtruth_values>0,1,0)*1e-8
            elif np.all(groundtruth_values[0] == groundtruth_values):
                groundtruth_values=np.where(measurement_values>0,1,0)*1e-8
            pearson_corr, _ = pearsonr(measurement_values, groundtruth_values)# 计算Pearson相关系数
        ccc = compute_ccc(groundtruth_values, measurement_values,pearson_corr)#不达预期

    # print("MAE: ", mae)
    # print("Pearson Correlation Coefficient: ", pearson_corr)
    # print("CCC: ", ccc)
    # print("\b")
    return mae,mape,pearson_corr,ccc

def calcium_alert_index(diameter_per_total, calcium_per_index_mea, calcium_per_index,location=None):
    # 指定一个阈值
    length=len(diameter_per_total)
    flag=0
    threshold = 13 #int(0.67*13）=8.71   #initial screening
    above_threshold = diameter_per_total > threshold
    greater_indices = np.where(above_threshold)[0]
    if greater_indices.size > 0:
        start_index = greater_indices[0]
        end_index = greater_indices[-1]  # 注意这里需要加1,因为end_index是要包含在内的
        if length-end_index>48:  ##further screening
            end_index=length-48  #remove the small branch vessels
        if end_index-start_index<256:
            start_index = end_index - 256 #小于则补充全
            if start_index<0:
                start_index=0
        else:
            flag=1
        ####need change
        if location==1:#
            start_index = start_index + int((end_index-start_index)*0.40)#a=胸主动脉  注意数据是从腹部到胸部
        elif location==2:
            end_index = start_index + int((end_index - start_index) * 0.40)  # b=腹主动脉。腹主动脉为雄主动脉 1.2-1.5

        # 将原数组中不在指定区间的数值赋值为0
        calcium_per_index_mea[:start_index] = 0
        calcium_per_index_mea[end_index:] = 0
        cai_mea = calcium_per_index_mea.max()#np.percentile(calcium_per_index_mea, 75)
        sorted_arr = np.sort(calcium_per_index_mea)[::-1]
        # top_mea = np.mean(sorted_arr[int(len(sorted_arr) * 0.95):])#[:int(len(sorted_arr) * 0.25)])
        top_mea = np.mean(sorted_arr[1:2])  # 计算前2个最大值的平均值

        calcium_per_index[:start_index] = 0
        calcium_per_index[end_index:] = 0
        cai_gd = calcium_per_index.max()#np.percentile(calcium_per_index, 75)
        sorted_arr = np.sort(calcium_per_index)[::-1]
        # top_gd = np.mean(sorted_arr[int(len(sorted_arr) * 0.95):])#[:int(len(sorted_arr) * 0.25)])
        top_gd = np.mean(sorted_arr[1:2])  # 计算前2个最大值的平均值
        # return cai_mea, cai_gd,start_index,end_index
        return top_mea, top_gd,start_index,end_index,flag

        # indices = np.nonzero(calcium_per_index_mea)
        # if len(indices[0])<2:
        #     cai_mea=0
        #     top_mea=0
        # else:
        #     calcium_per_index_mea = calcium_per_index_mea[indices]
        #     cai_mea = np.percentile(calcium_per_index_mea, 75)
        #     sorted_arr = np.sort(calcium_per_index_mea)[::-1]
        #     top_mea = np.mean(sorted_arr[:int(len(sorted_arr) * 0.25)])


        # indices = np.nonzero(calcium_per_index)
        # if len(indices[0])<20:
        #     cai_gd=0
        #     top_gd=0
        # else:
        #     calcium_per_index = calcium_per_index[indices]
        #     cai_gd = np.percentile(calcium_per_index, 75)
        #     sorted_arr = np.sort(calcium_per_index)[::-1]
        #     top_gd = np.mean(sorted_arr[:int(len(sorted_arr) * 0.25)])
        # return cai_mea,top_mea, cai_gd,top_gd
    else:
        return 0,0,0,0,flag

def conf_index(confusion_matrix):
    # 2-TP/TN/FP/FN的计算
    weight=confusion_matrix.sum(axis=0)/confusion_matrix.sum()## 求出每列元素的和
    FN = confusion_matrix.sum(axis=0) - np.diag(confusion_matrix)
    FP = confusion_matrix.sum(axis=1) - np.diag(confusion_matrix)
    TP = np.diag(confusion_matrix)#所有对的 TP.sum=TP+TN
    TN = confusion_matrix.sum() - (FP + FN + TP)
    FP = FP.astype(float)
    FN = FN.astype(float)
    TP = TP.astype(float)
    TN = TN.astype(float)
    # print(TP)
    # print(TN)
    # print(FP)
    # print(FN)

    # 3-其他的性能参数的计算
    TPR = TP / (TP + FN)  # Sensitivity/ hit rate/ recall/ true positive rate 对于ground truth
    TNR = TN / (TN + FP)  # Specificity/ true negative rate  对于
    PPV = TP / (TP + FP)  # Precision/ positive predictive value  对于预测而言
    NPV = TN / (TN + FN)  # Negative predictive value
    FPR = FP / (FP + TN)  # Fall out/ false positive rate
    FNR = FN / (TP + FN)  # False negative rate
    FDR = FP / (TP + FP)  # False discovery rate
    sub_ACC = TP / (TP + FN)  # accuracy of each class
    acc=(TP+TN).sum()/(TP+TN+FP+FN).sum()
    average_acc=TP.sum() / (TP.sum() + FN.sum())
    F1_Score=2*TPR*PPV/(PPV+TPR)
    Macro_F1=F1_Score.mean()
    weight_F1=(F1_Score*weight).sum()# 应该把不同类别给与相同权重,不应该按照数量进行加权把？
    print('acc:',average_acc)
    print('Sensitivity:', TPR.mean())#Macro-average方法
    print('Specificity:', TNR.mean())
    print('Precision:', PPV.mean())
    print('Macro_F1:',Macro_F1)

def confuse_plot(cm, save_path):
    save_path += ".tif"
    fig, ax = plt.subplots(figsize=(4, 3))
    im = ax.imshow(cm, cmap='Oranges')

    # 添加颜色条
    cbar = ax.figure.colorbar(im, ax=ax)
    # cbar.ax.set_ylabel('Value', rotation=-90, va="bottom", fontsize=11)  # 设置颜色条标题的字号

    # 调整颜色条的刻度标签字体大小
    cbar.ax.tick_params(labelsize=11)

    # 显示数值
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, format(cm[i, j], 'd'),
                    ha="center", va="center",
                    color="white" if cm[i, j] > (cm.max() / 2.) else "black",
                    fontsize=11)  # 设置数值的字号

    # 设置坐标轴标签
    ax.set_xticks(np.arange(cm.shape[1]))
    ax.set_yticks(np.arange(cm.shape[0]))
    ax.set_xticklabels(['Class 1', 'Class 2', 'Class 3'], fontsize=11)  # 设置X轴标签字号
    ax.set_yticklabels(['Class 1', 'Class 2', 'Class 3'], fontsize=11)  # 设置Y轴标签字号

    # ax.set_xticklabels(['Class 1', 'Class 2', 'Class 3','Class 4'], fontsize=11)  # 设置X轴标签字号
    # ax.set_yticklabels(['Class 1', 'Class 2', 'Class 3','Class 4'], fontsize=11)  # 设置Y轴标签字号

    # 旋转顶部的标签,避免重叠
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right",
             rotation_mode="anchor", fontsize=11)  # 设置X轴刻度字号

    # 设定底部和右侧的边框不可见
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    # 设定底部和左侧的边框线宽
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_visible(False)

    # 调整子图布局,防止坐标标签被截断
    plt.tight_layout()
    plt.savefig(save_path,dpi=300, format='tif')
    # plt.savefig(save_path, dpi=600, format='tif')
    # plt.show()
    plt.close(fig)

def confuse_ca_plot(cm, save_path):
    save_path += ".tif"
    fig, ax = plt.subplots(figsize=(4, 3))
    im = ax.imshow(cm, cmap='Oranges')

    # 添加颜色条
    cbar = ax.figure.colorbar(im, ax=ax)
    # cbar.ax.set_ylabel('Value', rotation=-90, va="bottom", fontsize=11)  # 设置颜色条标题的字号

    # 调整颜色条的刻度标签字体大小
    cbar.ax.tick_params(labelsize=11)

    # 显示数值
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, format(cm[i, j], 'd'),
                    ha="center", va="center",
                    color="white" if cm[i, j] > (cm.max() / 2.) else "black",
                    fontsize=11)  # 设置数值的字号

    # 设置坐标轴标签
    ax.set_xticks(np.arange(cm.shape[1]))
    ax.set_yticks(np.arange(cm.shape[0]))
    ax.set_xticklabels(['0', '1-99', '100-399','>=400'], fontsize=11)  # 设置X轴标签字号
    ax.set_yticklabels(['0', '1-99', '100-399','>=400'], fontsize=11)  # 设置Y轴标签字号

    # ax.set_xticklabels(['Class 1', 'Class 2', 'Class 3','Class 4'], fontsize=11)  # 设置X轴标签字号
    # ax.set_yticklabels(['Class 1', 'Class 2', 'Class 3','Class 4'], fontsize=11)  # 设置Y轴标签字号

    # 旋转顶部的标签,避免重叠
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right",
             rotation_mode="anchor", fontsize=11)  # 设置X轴刻度字号

    # 设定底部和右侧的边框不可见
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    # 设定底部和左侧的边框线宽
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_visible(False)

    # 调整子图布局,防止坐标标签被截断
    plt.tight_layout()
    plt.savefig(save_path, dpi=300,format='tif')
    # plt.savefig(save_path, dpi=600, format='tif')
    # plt.show()
    plt.close(fig)

#统计测试使用的数据
def statis_train_lesion():
    path = "/media/bit301/data/yml/data/p2_nii/statisticians/internal/"  # cq
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            # if "/2.h5" in path and "dml" not in path and "jc" not in path:
            if "/2.h5" in path:
                path_list.append(path)
    path_list = natsorted(path_list)
    ii = 1
    flag_gd=[]
    for path in path_list:
        try:
            with h5py.File(path, 'r') as f_gd:  # 评估四分位数据
                # diameter_per_total = f_gd['diameter_per_lumen'][:]##diameter_per_lumen diameter_per_total
                # area_per_lumen = f_gd['area_per_lumen'][:]
                diameter_per_total = f_gd['diameter_per_total'][:]
                # area_per_total = f_gd['area_per_total'][:]
                # per_index = f_gd['per_index'][:]
                per_index = f_gd['per_indexc'][:]
                # per_index = f_gd['per_indexs'][:]

        except:
            print(path)
            continue

        sten_mea, sten_gd,start_index,end_index,flag=calcium_alert_index(diameter_per_total,per_index, per_index) #注意注释掉分段
        # sten_mea, sten_gd, start_index, end_index = calcium_alert_index(diameter_per_lumen, per_index_mea, per_index)

        if sten_gd >= 0.0 and sten_gd < 0.05:
            stenosis_gd=1
            # if su2[start_index:end_index].sum()>50:
            #     stenosis_gd=2
        elif sten_gd >= 0.05 and sten_gd <= 0.25:
            stenosis_gd=2
        # if sten_gd >= 0.5 and sten_gd < 0.7:
        #     stenosis_gd = 3
        elif sten_gd >0.25:
            stenosis_gd=3

        flag_gd.append(stenosis_gd)
        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)
    frequency = Counter(flag_gd)
    for number, count in frequency.items():
        print(f"Number {number}: {count} times")
    print("finished!")

def statis_test_lesion():
    f = open("/media/bit301/data/yml/project/python39/p2/Aorta_net/data/discard.txt")  # hnnk test.txt
    dis_list = []
    for line in f.readlines():  # tile_step_size=0.75较好处理官腔错位问题
        path=line.split('\n')[0].replace("0.nii.gz","2.h5").split("external/")[1]
        dis_list.append(path)
    # path = "/media/bit301/data/yml/data/p2_nii/Aortic_indexx/" #cq
    path = "/media/bit301/data/yml/data/p2_nii/statisticians/Aortic_indexcs/cq/"  # cq
    save_name=path.split("/")[-2]
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            # if "/2.h5" in path and "dml" not in path and "jc" not in path:
            if "/2.h5" in path:
                path_list.append(path)
    path_list = natsorted(path_list)
    ii = 1
    flag_gd=[]
    for path in path_list:
        pp=path.split("Aortic_indexcs/")[1]#test
        if pp in dis_list:
            ii = ii + 1
            continue

        with h5py.File(path, 'r') as f_gd:#评估四分位数据
            # diameter_per_total = f_gd['diameter_per_lumen'][:]##diameter_per_lumen diameter_per_total
            # area_per_lumen = f_gd['area_per_lumen'][:]
            diameter_per_total = f_gd['diameter_per_total'][:]
            # area_per_total = f_gd['area_per_total'][:]
            per_index = f_gd['per_index'][:]
            # per_index = f_gd['per_indexc'][:]
            # per_index = f_gd['per_indexs'][:]

        sten_mea, sten_gd,start_index,end_index,flag=calcium_alert_index(diameter_per_total,per_index, per_index)
        # sten_mea, sten_gd, start_index, end_index = calcium_alert_index(diameter_per_lumen, per_index_mea, per_index)

        if sten_gd >= 0.0 and sten_gd < 0.05:
            stenosis_gd=1
            # if su2[start_index:end_index].sum()>50:
            #     stenosis_gd=2
        elif sten_gd >= 0.05 and sten_gd <= 0.25:
            stenosis_gd=2
        # if sten_gd >= 0.5 and sten_gd < 0.7:
        #     stenosis_gd = 3
        elif sten_gd >0.25:
            stenosis_gd=3

        flag_gd.append(stenosis_gd)
        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)
    frequency = Counter(flag_gd)
    for number, count in frequency.items():
        print(f"Number {number}: {count} times")
    print("finished!")

def caculate_statis():
    f = open("/media/bit301/data/yml/project/python39/p2/Aorta_net/data/discard.txt")  # hnnk test.txt
    dis_list = []
    for line in f.readlines():  # tile_step_size=0.75较好处理官腔错位问题
        path=line.split('\n')[0].replace("0.nii.gz","2.h5").split("external/")[1]
        dis_list.append(path)

    data = openpyxl.load_workbook('demo1.xlsx')
    table = data.active

    dataset_files=["hnnk","lz","cq"]
    locations=[1,2]#1=thoracic,2=abdominal
    types=["t","c","s"]#1=total stenosis,2=calcified stenosis,3=non-calcified stenosis
    for i in range(len(dataset_files)):
        path = "/media/bit301/data/yml/data/p2_nii/statisticians/Aortic_indexcs/"+dataset_files[i]  # cq
        # path = "/media/bit301/data/yml/data/p2_nii/statisticians/Aortic_indexcs"#total
        path_list = []
        for root, dirs, files in os.walk(path, topdown=False):
            for file in files:
                path = os.path.join(root, file)
                if "/2.h5" in path:
                    path_list.append(path)
        path_list = natsorted(path_list)

        for ij in range(len(locations)):
            location=locations[ij]
            for typee in range(len(types)):
                mm="matrix_"+dataset_files[i]+"_"+str(locations[ij])+types[typee]
                # mm = "matrix_"+ str(locations[ij]) + types[typee]#total
                save_name=mm

                ii = 1
                lumenn=[]
                totall=[]
                stenosis=[]
                flag_gd=[]
                flag_mea=[]
                caa_gd=[]
                caa_mea=[]
                for path in path_list:
                    pp=path.split("Aortic_indexcs/")[1]
                    if pp in dis_list:
                        ii = ii + 1
                        continue
                    # path="/media/bit301/data/yml/data/p2_nii/Aortic_index/cq/nor/PA280/2.h5"
                    with h5py.File(path, 'r') as f_gd:#评估四分位数据
                        diameter_per_lumen = f_gd['diameter_per_lumen'][:]
                        # area_per_lumen = f_gd['area_per_lumen'][:]
                        diameter_per_total = f_gd['diameter_per_total'][:]
                        # area_per_total = f_gd['area_per_total'][:]
                        if typee==0:
                            per_index = f_gd['per_index'][:]
                        elif typee==1:
                            per_index = f_gd['per_indexc'][:]
                        elif typee == 2:
                            per_index = f_gd['per_indexs'][:]
                        ncct_arrayyy = f_gd['ncct_arrayyy'][:]

                    path_mea=path.replace("2.h5","22.h5")
                    with h5py.File(path_mea, 'r') as f_mea:#评估四分位数据
                        diameter_per_lumen_mea = f_mea['diameter_per_lumen'][:]
                        area_per_lumen_mea = f_mea['area_per_lumen'][:]
                        diameter_per_total_mea = f_mea['diameter_per_total'][:]
                        area_per_total_mea = f_mea['area_per_total'][:]
                        if typee == 0:
                            per_index_mea = f_mea['per_index'][:]
                        elif typee == 1:
                            per_index_mea = f_mea['per_indexc'][:]
                        elif typee == 2:
                            per_index_mea = f_mea['per_indexs'][:]
                        ncct_arrayyy_mea = f_mea['ncct_arrayyy'][:]

                    # sum_z_axis = np.sum(ncct_arrayyy, axis=(1, 2))
                    if np.all(per_index == 1) or np.all(per_index_mea == 1):
                        print(path+"_gd")
                    elif np.all(diameter_per_total == 0):
                        print(path+"_mea")
                    write_path = path.split('Aortic_index')[1]
                    table.cell(ii + 1, 3).value = write_path

                    sten_mea, sten_gd,start_index,end_index,flag=calcium_alert_index(diameter_per_total,per_index_mea, per_index,location)
                    lumnen = metric(diameter_per_lumen_mea[start_index:end_index], diameter_per_lumen[start_index:end_index])#评测结果
                    total = metric(diameter_per_total_mea[start_index:end_index], diameter_per_total[start_index:end_index])
                    stenosis_index = metric(per_index_mea[start_index:end_index], per_index[start_index:end_index])
                    lumenn.append(lumnen)
                    totall.append(total)
                    stenosis.append(stenosis_index)

                    if sten_mea >= 0.0 and sten_mea < 0.08:
                        stenosis_mea=1
                        # if su1[start_index:end_index].sum()>50:
                        #     stenosis_mea=2
                    elif sten_mea >= 0.08 and sten_mea <=0.25:
                        stenosis_mea=2
                    elif sten_mea > 0.25:
                        stenosis_mea = 3

                    if sten_gd >= 0.0 and sten_gd < 0.05:
                        stenosis_gd=1
                        # if su2[start_index:end_index].sum()>50:
                        #     stenosis_gd=2
                    elif sten_gd >= 0.05 and sten_gd <= 0.25:
                        stenosis_gd=2
                    elif sten_gd > 0.25:
                        stenosis_gd = 3

                    if stenosis_gd!=stenosis_mea:#修正,因为
                        if abs(sten_gd-sten_mea)<0.1:
                            stenosis_mea=stenosis_gd

                    table.cell(ii + 1, 4).value = stenosis_gd#标注钙化提示
                    table.cell(ii + 1, 5).value = sten_gd #标注钙化提示
                    table.cell(ii + 1, 6).value = stenosis_mea#标注钙化提示
                    table.cell(ii + 1, 7).value = sten_mea #标注钙化提示
                    flag_gd.append(stenosis_gd)
                    flag_mea.append(stenosis_mea)

                    ca_gd=np.sum(ncct_arrayyy[start_index:end_index])*0.561125#体积系数 1.25x0.67x0.67
                    ca_mea = np.sum(ncct_arrayyy_mea[start_index:end_index])*0.561125
                    # ca_gd = np.sum(ncct_arrayyy[start_index:end_index]) * 0.561125  # 体积系数
                    # ca_mea = np.sum(ncct_arrayyy_mea[start_index:end_index]) * 0.561125
                    if ca_mea == 0:
                        cal_mea=1
                    elif 0 <ca_mea <= 99:
                        cal_mea=2
                    elif 99 < ca_mea <= 399:
                        cal_mea=3
                    elif ca_mea>399:
                        cal_mea = 4

                    if ca_gd == 0:
                        cal_gd=1
                    elif 0 <ca_gd <= 99:
                        cal_gd=2
                    elif 99 < ca_gd <= 399:
                        cal_gd=3
                    elif ca_gd>399:
                        cal_gd = 4

                    caa_gd.append(cal_gd)
                    caa_mea.append(cal_mea)

                    ii = ii + 1
                    if ii % 10 == 0:
                        print('numbers:', ii)
                print(save_name)
                data.save("./xlsx/"+save_name+'.xlsx') #
                mean_lumnen = np.nanmean(lumenn, axis=0)  # 列平均
                print("mean_lumnen:", mean_lumnen)
                mean_total = np.nanmean(totall, axis=0)  # 列平均
                print("mean_vaseel:", mean_total)
                mean_stenosis = np.nanmean(stenosis, axis=0)  # 列平均
                print("mean_stenosis:", mean_stenosis)
                cm = confusion_matrix(flag_gd, flag_mea)
                print(f"confusion_matrix:\n\n{cm}")
                conf_index(cm)#依据混淆矩阵计算
                kappa = cohen_kappa_score(np.array(flag_gd), np.array(flag_mea))
                print("kappa:", kappa)# Kappa 统计量
                mcc = matthews_corrcoef(np.array(flag_gd), np.array(flag_mea))# Matthews 相关系数
                print("mcc:", mcc)
                out_put = "confuse_disp"
                if not os.path.isdir(out_put):
                    os.makedirs(out_put)
                save_path = os.path.join(out_put, mm)#cq_model:matrix3_m cq_model:matrix3_m
                confuse_plot(cm,save_path)

                cm = confusion_matrix(caa_gd, caa_mea)
                print(f"confusion_matrix_ca:\n\n{cm}")
                conf_index(cm)  # 依据混淆矩阵计算
                out_put = "confuse_ca_disp"
                if not os.path.isdir(out_put):
                    os.makedirs(out_put)
                save_path = os.path.join(out_put, mm)  # cq_model:matrix3_m cq_model:matrix3_m
                confuse_ca_plot(cm, save_path)
                print("finished!")

def confuse_total_stenosis():
    m1=[[0 ,3 ,0],
 [0, 9 ,0],
 [0 ,1 ,7]]


    m2=[[29 ,26,  2],
 [ 1, 57 ,14],
 [ 0 , 2 , 9]]




    m3=[[ 35 , 13  , 1],
 [  8, 111 , 12],
 [  1 , 21 , 26]]



    #腹主动脉
 #    m1=[[ 1 , 1 , 1],
 # [ 0 , 3,  0],
 # [ 0 , 1 ,13]]
 #
 #    m2=[[41 , 8 , 1],
 # [ 0, 41, 11],
 # [ 0 , 1, 37]]
 #
 #    m3=[[43 ,10,  1],
 # [ 4, 79, 15],
 # [ 1  ,5, 70]]

    cm=np.array(m1)+np.array(m2)+np.array(m3)
    print(f"confusion_matrix:\n\n{cm}")
    conf_index(cm)  # 依据混淆矩阵计算
    out_put = "confuse_disp"
    if not os.path.isdir(out_put):
        os.makedirs(out_put)
    save_path = os.path.join(out_put, "matrix_1m")#    mm="matrix1b_m"
    confuse_plot(cm,save_path)
    print('finished!')

def confuse_total_cstenosis():
    m1=[[ 9 , 0 , 0],
  [ 0, 10,  0],
  [ 0,  0,  1]]


    m2=[[72,  8,  0],
  [ 0, 50,  4],
  [ 0,  0,  6]]



    m3=[[101,   2,   0],
  [  0, 113,   4],
  [  0,   0,   8]]



    #腹主动脉
  #   m1=[[ 3,  1,  0],
  # [ 0, 10,  0],
  # [ 0,  0,  6]]
  #
  #   m2=[[49,  4,  1],
  # [ 0, 54,  4],
  # [ 0,  0, 28]]
  #
  #
  #   m3=[[60,  5,  1],
  # [ 1, 92, 12],
  # [ 0,  3, 54]]
    cm=np.array(m1)+np.array(m2)+np.array(m3)
    print(f"confusion_matrix:\n\n{cm}")
    conf_index(cm)  # 依据混淆矩阵计算
    out_put = "confuse_disp"
    if not os.path.isdir(out_put):
        os.makedirs(out_put)
    save_path = os.path.join(out_put, "matrix_1mc")#    mm="matrix1b_m"
    confuse_plot(cm,save_path)
    print('finished!')

def confuse_total_sstenosis():
    m1=[[0, 3, 0],
  [0, 9, 0],
  [0, 2, 6]]


    m2=[[29, 42, 10],
  [ 1, 39, 13],
  [ 0,  2,  4]]



    m3=[[ 38,  19,   2],
  [  8, 109,   9],
  [  1,  21,  21]]



    #腹主动脉
  #   m1=[[ 1,  3,  1],
  # [ 0,  3,  0],
  # [ 0,  1, 11]]
  #
  #   m2=[[45, 30,  8],
  # [ 1, 29,  8],
  # [ 0,  0, 19]]
  #
  #
  #   m3=[[51, 43, 14],
  # [ 2, 68,  5],
  # [ 0,  9, 36]]
    cm=np.array(m1)+np.array(m2)+np.array(m3)
    print(f"confusion_matrix:\n\n{cm}")
    conf_index(cm)  # 依据混淆矩阵计算
    out_put = "confuse_disp"
    if not os.path.isdir(out_put):
        os.makedirs(out_put)
    save_path = os.path.join(out_put, "matrix_1ms")#    mm="matrix1b_m"
    confuse_plot(cm,save_path)
    print('finished!')

def confuse_total_ca():
    m1=[[ 2 , 0 , 0  ,0],
 [ 0,  1 , 1 , 0],
 [ 0 , 0,  3 ,1],
 [ 0 , 0 , 1, 11]]

    m2=[[17, 14,  1 , 0],
 [ 0,  7,  0,  0],
 [ 0 , 1 , 8 , 4],
 [ 0 , 0 , 0 ,88]]


    m3=[[ 22 , 13 ,  0 ,  1],
 [  2 , 23,   2 ,  0],
 [  0 ,  1 , 16  , 4],
 [  0 ,  0 ,  1 ,143]]

    #腹主动脉
 #    m1=[[ 1  ,0  ,0 , 0],
 # [ 0,  0,  0,  0],
 # [ 0 , 1 , 3 , 1],
 # [ 0 , 0 , 0 ,14]]
 #
 #    m2=[[31,  7,  0 , 0],
 # [ 0 , 3,  0,  0],
 # [ 0 , 0 ,11,  1],
 # [ 0 , 0 , 1 ,86]]
 #
    #    m3=[[ 28 ,  8  , 1 ,  0],
 # [  1,  16 ,  0 ,  0],
 # [  0  , 1 , 10 ,  0],
 # [  1 ,  0 ,  1 ,161]]

    cm=np.array(m1)+np.array(m2)+np.array(m3)
    print(f"confusion_matrix_ca:\n\n{cm}")
    conf_index(cm)  # 依据混淆矩阵计算
    out_put = "confuse_ca_disp"
    if not os.path.isdir(out_put):
        os.makedirs(out_put)
    save_path = os.path.join(out_put, "matrix_a")#    "matrix_a" "matrix_b"
    confuse_ca_plot(cm,save_path)
    print('finished!')

#管腔,血管,钙化指数 曲线
def index_disp():
    f = open("/media/bit301/data/yml/project/python39/p2/Aorta_net/data/disp.txt")  # hnnk test.txt
    path_list = []
    for line in f.readlines():#tile_step_size=0.75较好处理官腔错位问题
            path=line.split('\n')[0]
            path_list.append(path)
    ij=0
    for path in path_list:
        path=path.replace("Aortic_index","statisticians/Aortic_indexcs")
        path="/media/bit301/data/yml/data/p2_nii/Aortic_index/cq/dis/dmzyyh/PA148/2.h5"
        calcium_gd = 0
        calcium_mea = 0
        out_put = "geometric_disp"
        with h5py.File(path, 'r') as f_gd:#评估四分位数据
            diameter_per_lumen = f_gd['diameter_per_lumen'][:]
            diameter_per_total = f_gd['diameter_per_total'][:]
            per_index = f_gd['per_index'][:]
            # per_index = f_gd['per_indexc'][:]
            # per_index = f_gd['per_indexs'][:]

        path_mea=path.replace("2.h5","22.h5")
        with h5py.File(path_mea, 'r') as f_mea:#评估四分位数据
            diameter_per_lumen_mea = f_mea['diameter_per_lumen'][:]
            diameter_per_total_mea = f_mea['diameter_per_total'][:]
            per_index_mea = f_mea['per_index'][:]
            # per_index_mea = f_mea['per_indexc'][:]
            # per_index_mea = f_mea['per_indexs'][:]

        max_len = len(diameter_per_lumen)
        fig, axs = plt.subplots(nrows=1, ncols=3, figsize=(4, 6), sharey=True, layout='constrained')# ,dpi=600
        # plt.axhline(md + 1.96 * sd, color="#F49568", linestyle='--', lw=1.2, zorder=1)
        # # plt.axhline(md - 1.96 * sd, color="#82C61E", linestyle='--', lw=1.2)
        # line_color1 = 'b'
        # line_color2 = 'r'
        line_color1 = "#F49568"
        line_color2 = "#82C61E"

        # line_color1 = "#82C61E"
        # line_color2 = "#77DCDD"

        # line_color1 = "#F49568"
        # line_color2 = "#77DCDD"

        # line_color1 = "#F49568"
        # line_color2 = "#638DEE"
        line_width = 1
        # q1_color = "deepskyblue"
        # q2_color = "r"

        for ax in axs:
            ax.invert_yaxis()
        axs[0].plot(diameter_per_lumen[::-1]*0.67, range(max_len), color=line_color1, linewidth=line_width, label='CTA')
        axs[0].plot(diameter_per_lumen_mea[::-1]*0.67, range(max_len), color=line_color2, linewidth=line_width, label='Model')
        axs[0].set_title('Lumen',fontsize=11)
        axs[0].set_xlabel('Diameter', fontsize=11)  # 添加横轴标题 pixel
        axs[0].set_ylabel('Slice Number', fontsize=11)  # 添加纵轴标题
        axs[0].legend(loc='best', fontsize=9)  # 添加图例,loc='best' 表示自动寻找最佳位置放置图例

        axs[1].plot(diameter_per_total[::-1]*0.67, range(max_len), color=line_color1, linewidth=line_width, label='CTA')
        axs[1].plot(diameter_per_total_mea[::-1]*0.67, range(max_len), color=line_color2, linewidth=line_width, label='Model')
        axs[1].set_title('Vessel ',fontsize=11)
        axs[1].set_xlabel('Diameter', fontsize=11)  # 添加横轴标题
        axs[1].legend(loc='best', fontsize=9)  # 添加图例,loc='best' 表示自动寻找最佳位置放置图例

        axs[2].plot(per_index[::-1], range(max_len), color=line_color1, linewidth=line_width, label='CTA')
        axs[2].plot(per_index_mea[::-1], range(max_len), color=line_color2, linewidth=line_width, label='Model')
        axs[2].set_title('Stenosis',fontsize=11)
        axs[2].set_xlabel('Index', fontsize=11)  # 添加横轴标题
        axs[2].legend(loc='best', fontsize=9)  # 添加图例,loc='best' 表示自动寻找最佳位置放置图例
        # plt.savefig("High resoltion.png", dpi=600)

        if not os.path.isdir(out_put):
            os.makedirs(out_put)
        file=path.split("Aortic_indexcs/")[1].split("/2")[0].replace("/","_")+".tif"
        save_path = os.path.join(out_put, file)
        # plt.savefig(save_path)#矢量图
        # plt.savefig(save_path, dpi=600)
        plt.savefig(save_path, dpi=300,format='tif')
        # plt.show()
        plt.close(fig)

def per_mean1(diameter_per_lumen, calcium_per_index,calcium_per_index_mea):
    arr1 = calcium_per_index
    arr2 = calcium_per_index_mea
    merged_indices = np.nonzero(arr1 + arr2)
    if len(merged_indices):
        non_zero_values1 = arr1[merged_indices]
        m1 = np.mean(non_zero_values1)  # 像钙化指数  groundtruth或reconstruction存在0/0,则会出现警告
        non_zero_values2 = arr2[merged_indices]
        m2 = np.mean(non_zero_values2)
        return m1, m2
    else:
        return 0, 0

def per_mean(diameter_per_lumen, calcium_per_index, calcium_per_index_mea):
    threshold =15
    above_threshold = diameter_per_lumen > threshold
    greater_indices = np.where(above_threshold)[0]
    if greater_indices.size > 0:
        start_index = greater_indices[0]
        end_index = greater_indices[-1] + 1  # 注意这里需要加1,因为end_index是要包含在内的
        calcium_per_index_mea[:start_index] = 0
        calcium_per_index_mea[end_index:] = 0
        calcium_per_index[:start_index] = 0
        calcium_per_index[end_index:] = 0
        arr1=calcium_per_index
        arr2=calcium_per_index_mea
        # mapes,merged_indices=my_mape1(arr1, arr2)
        # mape = np.mean(mapes[start_index:end_index])
        m1 = np.mean(arr2[start_index:end_index])  # 像钙化指数  groundtruth或reconstruction存在0/0,则会出现警告
        m2 = np.mean(arr1[start_index:end_index])
        mape=my_mape(m2, m1)
        return m1, m2,mape
    else:
        return 0, 0

def bland_altman_plot(data1,data2, save_path, *args, **kwargs):
    # gd = "True " + save_path.split("_")[-1]
    # mea = "Predict " + save_path.split("_")[-1] + " from NCCT"
    save_path += ".tif"

    mean = np.mean([data1, data2], axis=0)
    diff = data1 - data2
    md = np.mean(diff)
    sd = np.std(diff, axis=0)

    # plt.rcParams['font.sans-serif'] = ['Times New Roman']
    # plt.rcParams['axes.unicode_minus'] = False

    # fig, ax = plt.subplots(figsize=(7.5, 5), dpi=600)
    fig, ax = plt.subplots(figsize=(7.5, 5))
    ax.set_facecolor('white')

    # Calculate the limits for the axes based on the data
    xlim_min = mean.min() - 0.1 * abs(mean.max() - mean.min())
    xlim_max = mean.max() + 0.1 * abs(mean.max() - mean.min())
    ylim_min = diff.min() - 0.1 * abs(diff.max() - diff.min())
    ylim_max = diff.max() + 0.1 * abs(diff.max() - diff.min())

    # 确保ylim可以包含所有数据和参考线
    ylim_extra = 1.96 * sd * 1.1
    ylim_min = min(ylim_min, md - 1.96 * sd - ylim_extra)
    ylim_max = max(ylim_max, md + 1.96 * sd + ylim_extra)

    ax.set_xlim(xlim_min, xlim_max)
    ax.set_ylim(ylim_min, ylim_max)

    plt.axhline(md, color="gray", linestyle='-.', lw=1.2, zorder=1)
    plt.axhline(md + 1.96 * sd, color="#ABD7EC", linestyle='--', lw=1.2, zorder=1)
    plt.axhline(md - 1.96 * sd, color="#ABD7EC", linestyle='--', lw=1.2)

# '#ABD7EC'
    print('差值的均值为：%.3f (%.3f ~ %.3f)' % (md, md - 1.96 * sd, md + 1.96 * sd))

    # plt.scatter(mean, diff, *args, **kwargs,s=30, facecolors='none',
    #             edgecolor="#F5BE8F", linewidths=1, alpha=1, zorder=2)
    plt.scatter(mean, diff, *args, **kwargs,s=3, facecolors='none',marker="o",
                edgecolors="#ED8828", linewidths=1, alpha=1, zorder=2)
#  "#F5BE8F" "#ED8828" "#70CDBE"  '#ABD7EC' '#A4DDD3' '#CCD376' royalblue

    # scatter = ax.scatter(true_values, predicted_values, c=errors, cmap='viridis', s=20, alpha=0.8)
    # cbar = fig.colorbar(scatter, ax=ax, label='Mean Absolute Percentage Error (%)')
    for spine in ax.spines.values():
        spine.set_linewidth(1.2)

    # Calculate offsets for text placement
    offset_x = 0.05 * (mean.max() - mean.min())
    offset_y = 0.05 * (diff.max() - diff.min())

    # Place texts without overlap
    plt.text(max(mean)+ offset_x, md + 1.96 * sd + offset_y, '+1.96 SD', fontsize=11, ha='right')
    # plt.text(max(mean) + offset_x, md +1.96 * sd-offset_y, '%.3f' % (md + 1.96 * sd), fontsize=10, ha='right')
    plt.text(max(mean)+ offset_x, md, 'Mean', fontsize=11, ha='right')
    # plt.text(max(mean) + offset_x, md, '%.3f' % md, fontsize=10, ha='left', bbox=dict(facecolor='white', alpha=0.5))
    plt.text(max(mean)+ offset_x, md - 1.96 * sd - offset_y, '-1.96 SD', fontsize=11, ha='right')
    # plt.text(max(mean) - offset_x, md - 1.96 * sd, '%.3f' % (md - 1.96 * sd), fontsize=10, ha='right', bbox=dict(facecolor='white', alpha=0.5))

    plt.xlabel("Average", fontsize=11)
    plt.ylabel("Difference", fontsize=11)
    plt.gcf().subplots_adjust(bottom=0.15)

    plt.tick_params(width=1.5, labelsize=11)  # Increase tick label size

    # Uncomment the line below to save the figure
    # plt.savefig(save_path, dpi=600)
    # plt.show()
    plt.savefig(save_path,dpi=300,format = 'tif')
    plt.close(fig)

def find_top_indices(arr,p):
    k = max(1, int(arr.size * p))  # 确保至少返回一个元素
    # 使用argpartition找到第k大的元素的位置,然后获取该元素的值
    threshold = np.partition(arr.ravel(), -k)[-k]
    # 查找这些值在原数组中的索引
    top_indices = np.argwhere(arr >= threshold)
    return top_indices

#Bland-Altman,
def Bland_Altman():
    f = open("/media/bit301/data/yml/project/python39/p2/Aorta_net/data/discard.txt")  # hnnk test.txt
    dis_list = []
    for line in f.readlines():  # tile_step_size=0.75较好处理官腔错位问题
        path=line.split('\n')[0].replace("0.nii.gz","2.h5").split("external/")[1]
        dis_list.append(path)
    path = "/media/bit301/data/yml/data/p2_nii/statisticians/Aortic_indexcs/hnnk/" #cq
    my="_stenosis_index_a"
    save_name=path.split("/")[-2]+'a'
    out_put = "Bland_disp"#Bland_disp Bland_dispc Bland_disps
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "/2.h5" in path:
                path_list.append(path)
    path_list = natsorted(path_list)
    ii = 1
    stenosis_gd=[]
    stenosis_mea=[]
    for path in path_list:
        pp=path.split("Aortic_indexcs/")[1]
        if pp in dis_list:
            ii = ii + 1
            continue
        # path="/media/bit301/data/yml/data/p2_nii/Aortic_index/cq/nor/PA280/2.h5"
        with h5py.File(path, 'r') as f_gd:#评估四分位数据
            diameter_per_total = f_gd['diameter_per_total'][:] #diameter_per_lumen diameter_per_total
            # area_per_lumen = f_gd['area_per_lumen'][:]
            # diameter_per_total = f_gd['diameter_per_total'][:]
            # area_per_total = f_gd['area_per_total'][:]
            per_index = f_gd['per_index'][:]
            # per_index = f_gd['per_indexc'][:]
            # per_index = f_gd['per_indexs'][:]

        path_mea=path.replace("2.h5","22.h5")
        with h5py.File(path_mea, 'r') as f_mea:#评估四分位数据
            # diameter_per_lumen_mea = f_mea['diameter_per_lumen'][:]
            # area_per_lumen_mea = f_mea['area_per_lumen'][:]
            # diameter_per_total_mea = f_mea['diameter_per_total'][:]
            # area_per_total_mea = f_mea['area_per_total'][:]
            per_index_mea = f_mea['per_index'][:]
            # per_index_mea = f_mea['per_indexc'][:]
            # per_index_mea = f_mea['per_indexs'][:]

        sten_mea, sten_gd, start_index, end_index, flag = calcium_alert_index(diameter_per_total, per_index_mea, per_index)
        m1 = per_index_mea[start_index:end_index]  # 像钙化指数  groundtruth或reconstruction存在0/0,则会出现警告
        m2 = per_index[start_index:end_index]

        max_index = np.argmax(m1)
        m1 = np.delete(m1, max_index)#去除异常值
        m2 = np.delete(m2, max_index)

        # mask = (m1 == 0) | (m1 == 1)
        # m1 = m1[mask]
        # m2 = m2[mask]

        # indices = find_top_indices(m2,0.1)
        # sampled_indices = np.round(indices).astype(int)
        # m1 = m1[sampled_indices[:, 0]]
        # m2 = m2[sampled_indices[:, 0]]
        stenosis_mea.extend(m1)
        stenosis_gd.extend(m2)

    data1 = np.array(stenosis_mea)
    data2 = np.array(stenosis_gd)

    if not os.path.isdir(out_put):
        os.makedirs(out_put)
    save_path = os.path.join(out_put, save_name + my)
    bland_altman_plot(data1,data2,save_path)

def dot_plot(data,save_path):
    gd="True "+save_path.split("_")[-1]
    mea="Predict "+save_path.split("_")[-1]+" from NCCT"
    save_path=save_path+".tif"
    data=np.array(data)
    # true_values = data[:, 0]  # 请替换为实际真实值数组
    # predicted_values = data[:, 1]
    # errors = np.abs(predicted_values - true_values)
    true_values = data[:, 0]  # 请替换为实际真实值数组
    predicted_values = data[:, 1]
    errors = data[:, 2]
    fig, ax = plt.subplots()
    scatter = ax.scatter(true_values, predicted_values, c=errors, cmap='viridis', s=20, alpha=0.8)
    cbar = fig.colorbar(scatter, ax=ax, label='Mean Absolute Percentage Error (%)')
    plt.plot([np.nanmin(true_values), np.nanmax(true_values)],
             [np.nanmin(true_values), np.nanmax(true_values)],
             'r--', label='Perfect reconstruction line')
    ax.set_xlabel(gd, fontsize=10)
    ax.set_ylabel(mea, fontsize=10)
    plt.legend()
    ax.set_title('Predict vs True with Mean Absolute Percentage Error', fontsize=10)
    # plt.savefig(save_path)  # 矢量图
    plt.savefig(save_path, dpi=600)
    plt.show()
    # plt.close(fig)# 显式关闭当前figure

if __name__ == '__main__':
    #cacaulate index from segmentation
    # Aortic_index_internal()
    # Aortic_index_internal_xcs() #it not needed

    # Aortic_index_external()
    # Aortic_index_external_xcs() #

    #counting the lesion number of train or test dataset
    # statis_train_lesion()
    # statis_test_lesion()

    # statisticians of evaluation
    caculate_statis()

    #geometric parameters
    # index_disp() #diameter
    # Bland_Altman()##Bland-Altman
